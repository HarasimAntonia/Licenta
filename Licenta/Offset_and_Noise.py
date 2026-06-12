import serial
import time
import statistics
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# Serial communication and sample number
PORT = 'COM10'
BAUDRATE = 115200
SAMPLE = 5

# Number of measurements and name of the Excel file
READINGS = 100
EXCEL_FILE = f'../Offset_and_Noise_SAMPLE{SAMPLE}.xlsx'

def create_offset_graphic(sheet, readings, position):
    """ Generate the graphic for Offset (axis X, Y, Z) """

    chart = LineChart()
    chart.title = "Offset Analysis"
    chart.style = 13
    chart.y_axis.title = "Magnetic Field (mT)"
    chart.x_axis.title = "Number of measurement"
    chart.width = 20
    chart.height = 10

    # Limits on Y axis
    chart.y_axis.scaling.min = -3.0
    chart.y_axis.scaling.max = 3.0
    chart.y_axis.majorUnit = 1.0

    # Read measured data: X, Y, Z (Columns B, C, D)
    data_measured = Reference(sheet, min_col=2, min_row=1, max_col=4, max_row=readings + 1)
    chart.add_data(data_measured, titles_from_data=True)

    # Read limit data (Columns AA and AB)
    data_limits = Reference(sheet, min_col=27, min_row=1, max_col=28, max_row=readings + 1)
    chart.add_data(data_limits, titles_from_data=True)

    # Legend
    if len(chart.series) >= 5:
        chart.series[0].title = SeriesLabel(v="Offset X")
        chart.series[1].title = SeriesLabel(v="Offset Y")
        chart.series[2].title = SeriesLabel(v="Offset Z")

        max_line = chart.series[3]
        max_line.title = SeriesLabel(v="Maximum")
        max_line.graphicalProperties.line.solidFill = "FF0000"  # Red
        max_line.graphicalProperties.line.dashStyle = "sysDot"
        max_line.graphicalProperties.line.width = 25000

        min_line = chart.series[4]
        min_line.title = SeriesLabel(v="Minimum")
        min_line.graphicalProperties.line.solidFill = "FF0000"  # Red
        min_line.graphicalProperties.line.dashStyle = "sysDot"
        min_line.graphicalProperties.line.width = 25000

    sheet.add_chart(chart, position)

def create_noise_graphic(sheet, col_start, title, position, readings):
    """ Generate an individual Noise graphic for each axis"""

    chart = LineChart()
    chart.title = title
    chart.style = 13
    chart.y_axis.title = "Noise (uT)"
    chart.x_axis.title = "Number of measurement"
    chart.width = 20
    chart.height = 10
   # chart.y_axis.scaling.min = 0

    # Read data from the helper columns
    ref = Reference(sheet, min_col=col_start, min_row=1, max_col=col_start + 1, max_row=readings + 1)
    chart.add_data(ref, titles_from_data=True)

    # Legend
    if len(chart.series) >= 2:
        chart.series[0].title = SeriesLabel(v="Calculated Noise")

        max_line = chart.series[1]
        max_line.title = SeriesLabel(v="Maximum")
        max_line.graphicalProperties.line.solidFill = "FF0000"
        max_line.graphicalProperties.line.dashStyle = "sysDot"
        max_line.graphicalProperties.line.width = 25000

    sheet.add_chart(chart, position)

def offset_and_noise():
    """ Read X, Y and Z values from the serial and perform average for offset and standard deviation for noise """
    arduino = None
    try:
        print(f"[{PORT}] Connecting to Arduino.")
        arduino = serial.Serial(PORT, BAUDRATE, timeout=3.0)
        time.sleep(2)

        print(f"[{PORT}] Supply the sensor from the Arduino.")
        arduino.write('power_arduino\n'.encode('utf-8'))
        time.sleep(1)

        print("Sending command to Arduino to read the data.")
        arduino.write('reading\n'.encode('utf-8'))

        x_sum = 0.0
        y_sum = 0.0
        z_sum = 0.0
        active_readings = 0

        # Lists to store individual readings for noise calculation
        x_list = []
        y_list = []
        z_list = []

        # Excel initialization
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Offset data"

        # Table header
        sheet.append(["No.", "X (mT)", "Y (mT)", "Z (mT)"])

        print(f"Reading {READINGS} values.")
        while active_readings < READINGS:
            # Reading the line from arduino and decode it.
            line_before = arduino.readline()
            if not line_before:
                print("\n[ERROR] Timeout.")
                break
            line_after = line_before.decode('utf-8').strip()
            elements = line_after.split()
            if len(elements) == 4:
                x = float(elements[0])
                y = float(elements[1])
                z = float(elements[2])

                x_sum += x
                y_sum += y
                z_sum += z

                # Save each reading to its corresponding list
                x_list.append(x)
                y_list.append(y)
                z_list.append(z)

                active_readings += 1

                # Add data to table
                sheet.append([active_readings, x, y, z])
                print(f"[{active_readings:03d}/{READINGS}] X: {x:.2f} | Y: {y:.2f} | Z: {z:.2f}")

        # Average, Noise, and Graphic
        if active_readings == READINGS:
            # Calculate Offset (Mean)
            offset_x = round(x_sum / READINGS, 3)
            offset_y = round(y_sum / READINGS, 3)
            offset_z = round(z_sum / READINGS, 3)

            # Calculate Noise (Standard Deviation) and convert to uT
            noise_x = round(statistics.stdev(x_list) * 1000, 3)
            noise_y = round(statistics.stdev(y_list) * 1000, 3)
            noise_z = round(statistics.stdev(z_list) * 1000, 3)

            # Helper data moved far to the right
            sheet.cell(row=1, column=27, value="Max Lim (mT)")
            sheet.cell(row=1, column=28, value="Min Lim (mT)")
            sheet.cell(row=1, column=29, value="Noise X (uT)")
            sheet.cell(row=1, column=30, value="Limit X (uT)")
            sheet.cell(row=1, column=31, value="Noise Y (uT)")
            sheet.cell(row=1, column=32, value="Limit Y (uT)")
            sheet.cell(row=1, column=33, value="Noise Z (uT)")
            sheet.cell(row=1, column=34, value="Limit Z (uT)")

            # Write 100 rows of limits for offset and noise
            for r in range(2, READINGS + 2):
                sheet.cell(row=r, column=27, value=0.5)     # Offset Max Limit
                sheet.cell(row=r, column=28, value=-0.5)    # Offset Min Limit
                sheet.cell(row=r, column=29, value=noise_x) # Noise X
                sheet.cell(row=r, column=30, value=250.0)   # Noise Limit X
                sheet.cell(row=r, column=31, value=noise_y) # Noise Y
                sheet.cell(row=r, column=32, value=250.0)   # Noise Limit Y
                sheet.cell(row=r, column=33, value=noise_z) # Noise Z
                sheet.cell(row=r, column=34, value=173.0)   # Noise Limit Z

            # Saving the final values for offset and noise in different cells
            sheet['F1'], sheet['G1'], sheet['H1'] = "Offset X (mT)", "Offset Y (mT)", "Offset Z (mT)"
            sheet['F2'], sheet['G2'], sheet['H2'] = offset_x, offset_y, offset_z
            sheet['F4'], sheet['G4'], sheet['H4'] = "Noise X (uT)", "Noise Y (uT)", "Noise Z (uT)"
            sheet['F5'], sheet['G5'], sheet['H5'] = noise_x, noise_y, noise_z

            # Create offset graphic starting with K2 cell and noise graphic starting with K23
            create_offset_graphic(sheet, READINGS, position="K2")
            create_noise_graphic(sheet, 29, "Noise X", "K23", READINGS)
            create_noise_graphic(sheet, 31, "Noise Y", "K44", READINGS)
            create_noise_graphic(sheet, 33, "Noise Z", "K65", READINGS)

            # Saving in Excel
            try:
                wb.save(EXCEL_FILE)
                print(f"\n[SUCCESS] Excel file saved successfully as '{EXCEL_FILE}'.")
            except PermissionError:
                print(f"\n[ERROR] Cannot save '{EXCEL_FILE}'.")

    except Exception as e:
        print(f"\n[ERROR] Something went wrong :) {e}")
    finally:
        print("\nClosing the communication.")
        arduino.close()

if __name__ == '__main__':
    offset_and_noise()