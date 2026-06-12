import serial
import time
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from Licenta.drivers.SourceMeter import Keithley2400

# SMU address
SMU_ADDR = 'GPIB0::21::INSTR'

# Serial communication and sample number
PORT = 'COM10'
BAUDRATE = 115200
SAMPLE = 1

# Limits and name of the Excel file
EXCEL_FILE = f'../Power_Down_Current_SAMPLE{SAMPLE}.xlsx'
PD_LIMIT_NA = 1000.0
PD_TYPICAL_NA = 500.0

def create_power_down_graphic(sheet, readings, position):
    """ Generate the graphic for Power Down Current """
    chart = LineChart()
    chart.title = "Power Down Current Analysis"
    chart.style = 13
    chart.y_axis.title = "Current (nA)"
    chart.x_axis.title = "Number of measurement"
    chart.width = 20
    chart.height = 10

    # Limits on Y axis
    chart.y_axis.scaling.min = 0.0
    chart.y_axis.scaling.max = 1200.0
    chart.y_axis.majorUnit = 200.0

    # Read measured data (Column B)
    data_measured = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=readings + 1)
    chart.add_data(data_measured, titles_from_data=True)

    # Read limit data (Column C)
    data_limit = Reference(sheet, min_col=25, min_row=1, max_col=25, max_row=readings + 1)
    chart.add_data(data_limit, titles_from_data=True)

    # Read typical line (Column D)
    data_typical = Reference(sheet, min_col=26, min_row=1, max_col=26, max_row=readings + 1)
    chart.add_data(data_typical, titles_from_data=True)

    # Legend
    if len(chart.series) >= 3:
        chart.series[0].title = SeriesLabel(v="Power Down Current")

        max_line = chart.series[1]
        max_line.title = SeriesLabel(v="Maximum")
        max_line.graphicalProperties.line.solidFill = "FF0000"  # Red
        max_line.graphicalProperties.line.dashStyle = "sysDot"
        max_line.graphicalProperties.line.width = 25000

        typical_line = chart.series[2]
        typical_line.title = SeriesLabel(v="Typical")
        typical_line.graphicalProperties.line.solidFill = "00B050"  # Green
        typical_line.graphicalProperties.line.dashStyle = "sysDot"
        typical_line.graphicalProperties.line.width = 25000

    sheet.add_chart(chart, position)

def power_down_current_test():
    smu = Keithley2400()
    arduino = None
    try:
        print(f"[{PORT}] Connecting to Arduino.")
        arduino = serial.Serial(PORT, BAUDRATE, timeout=2.0)
        time.sleep(2)

        print(f"[{PORT}] Reset sensor.")
        arduino.write('reset_sensor\n'.encode('utf-8'))

        print(f"[SMU] Initializing SMU.")
        idn = smu.Initialize(visa_addr=SMU_ADDR, IdQuery=True, Reset=True)
        print(f"Connected to: {idn}")

        print(f"[{PORT}] Supply the sensor from the SourceMeter.")
        arduino.write('power_smu\n'.encode('utf-8'))
        time.sleep(0.5)

        # Excel initialization
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Power Down Current Data"

        # Table header
        sheet.append(["No.", "Power Down Current (nA)"])
        sheet.cell(row=1, column=25, value="Max Limit (nA)")
        sheet.cell(row=1, column=26, value="Typical (nA)")

        for i in range(1, 21):
            smu.configure_sense_function(mode="CURR")
            smu.configure_source_function(mode="VOLT")
            smu.configure_source_range_value(range_value=20.0)
            smu.configure_compliance(compliance=0.01)
            smu.configure_source_level(level=5.0)
            smu.configure_sense_range(range_value=0.01)
            smu.configure_output('ON')
            time.sleep(1)

            print(f"[{PORT}] Sending command to send sensor into power down.")
            arduino.write('power_down\n'.encode('utf-8'))
            time.sleep(0.5)

            print(f"[{PORT}] Disconnect SDA and SCL.")
            arduino.write('disconnect_SDA_SCL\n'.encode('utf-8'))
            time.sleep(0.5)

            smu.configure_sense_range(range_value=1e-6)
            time.sleep(1)

            power_down_current = smu.read_measurement()
            current_ua = power_down_current * 1000000000
            print(f"[{i}] Power down current: {current_ua:.4f} nA")

            sheet.append([i, current_ua])
            sheet.cell(row=i + 1, column=25, value=PD_LIMIT_NA)
            sheet.cell(row=i + 1, column=26, value=PD_TYPICAL_NA)

            print(f"[{PORT}] Connect SDA and SCL.\n")
            arduino.write('connect_SDA_SCL\n'.encode('utf-8'))
            time.sleep(0.5)

        create_power_down_graphic(sheet, 5, position="D2")

        # Saving in Excel
        try:
            wb.save(EXCEL_FILE)
            print(f"\n[SUCCESS] Excel file saved successfully as '{EXCEL_FILE}'.")
        except PermissionError:
            print(f"\n[ERROR] Cannot save '{EXCEL_FILE}'.")

    except Exception as e:
        print(f"\n[ERROR] Something went wrong :) {e}")
    finally:
        print("\nClosing the communication.")
        smu.configure_output('OFF')
        smu.Close()
        arduino.close()

if __name__ == "__main__":
    power_down_current_test()